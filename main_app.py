# -*- coding: utf-8 -*-
"""
DATA & FLOW PREDICTOR
Software Profesional para Aseguramiento de Flujo
Instituto Mexicano del Petróleo - Posgrado en Ingeniería de Hidrocarburos

Autor: IMP / Módulo de IA
"""

import sys, os

# ── Compatibilidad PyInstaller ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR   = os.path.join(BASE_DIR, "data")
MODELS_DIR = os.path.join(os.path.expanduser("~"), ".data_flow_predictor", "models")
os.makedirs(MODELS_DIR, exist_ok=True)

# ── Imports ─────────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns
import tensorflow as tf
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import RobustScaler, StandardScaler
from sklearn.metrics import r2_score, mean_absolute_percentage_error, mean_squared_error
import statsmodels.formula.api as smf
import joblib
import io
import base64
import warnings
warnings.filterwarnings("ignore")

# ── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Data & Flow Predictor | IMP",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS Corporativo IMP ───────────────────────────────────────────────────────
IMP_GREEN  = "#006847"
IMP_LIGHT  = "#e8f5ee"
IMP_GOLD   = "#C9A84C"
IMP_DANGER = "#c0392b"
IMP_WARN   = "#e67e22"

st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;600;700;900&family=Source+Code+Pro:wght@400;600&display=swap');

  html, body, [class*="css"] {{
      font-family: 'Montserrat', sans-serif;
      background: #f4f6f5;
  }}

  /* ── Sidebar ── */
  section[data-testid="stSidebar"] {{
      background: linear-gradient(180deg, {IMP_GREEN} 0%, #004d33 100%);
  }}
  section[data-testid="stSidebar"] * {{
      color: white !important;
  }}
  section[data-testid="stSidebar"] .stRadio > label {{
      background: rgba(255,255,255,0.08);
      border-radius: 8px;
      padding: 6px 12px;
      margin: 2px 0;
      transition: background 0.2s;
  }}
  section[data-testid="stSidebar"] .stRadio > label:hover {{
      background: rgba(255,255,255,0.18);
  }}

  /* ── Header strip ── */
  .imp-header {{
      background: linear-gradient(90deg, {IMP_GREEN} 0%, #004d33 60%, #003d28 100%);
      border-radius: 12px;
      padding: 22px 30px;
      display: flex;
      align-items: center;
      gap: 20px;
      margin-bottom: 22px;
      box-shadow: 0 4px 18px rgba(0,104,71,0.25);
  }}
  .imp-header h1 {{
      color: white;
      font-size: 1.75rem;
      font-weight: 900;
      margin: 0;
      letter-spacing: -0.5px;
  }}
  .imp-header span {{
      color: {IMP_GOLD};
      font-weight: 300;
      font-size: 0.9rem;
  }}

  /* ── KPI Cards ── */
  .kpi-card {{
      background: white;
      border-left: 5px solid {IMP_GREEN};
      border-radius: 10px;
      padding: 16px 20px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
  }}
  .kpi-card.warn  {{ border-left-color: {IMP_WARN};   }}
  .kpi-card.danger{{ border-left-color: {IMP_DANGER}; }}
  .kpi-card.gold  {{ border-left-color: {IMP_GOLD};   }}
  .kpi-val  {{ font-size: 2rem; font-weight: 900; color: {IMP_GREEN}; }}
  .kpi-lbl  {{ font-size: 0.78rem; color: #666; text-transform: uppercase; letter-spacing: 1px; }}

  /* ── Section titles ── */
  .sec-title {{
      font-size: 1.05rem;
      font-weight: 700;
      color: {IMP_GREEN};
      border-bottom: 2px solid {IMP_GREEN};
      padding-bottom: 4px;
      margin: 20px 0 14px;
  }}

  /* ── Status badges ── */
  .badge-safe   {{ background:#d4efdf; color:#1e8449; padding:4px 12px; border-radius:20px; font-weight:700; font-size:0.82rem; }}
  .badge-warn   {{ background:#fdebd0; color:#a04000; padding:4px 12px; border-radius:20px; font-weight:700; font-size:0.82rem; }}
  .badge-danger {{ background:#fadbd8; color:#922b21; padding:4px 12px; border-radius:20px; font-weight:700; font-size:0.82rem; }}

  /* ── Metric card override ── */
  [data-testid="metric-container"] {{
      background: white;
      border-radius: 10px;
      padding: 12px 16px;
      box-shadow: 0 2px 10px rgba(0,0,0,0.06);
  }}

  /* ── Tables ── */
  .dataframe {{ font-size: 0.82rem !important; }}

  /* ── Buttons ── */
  .stButton > button {{
      background: {IMP_GREEN};
      color: white;
      border: none;
      border-radius: 8px;
      font-weight: 700;
      padding: 0.5rem 1.4rem;
      transition: background 0.2s, transform 0.1s;
  }}
  .stButton > button:hover {{
      background: #004d33;
      transform: translateY(-1px);
  }}

  /* ── Expander ── */
  .streamlit-expanderHeader {{
      background: {IMP_LIGHT} !important;
      border-radius: 8px !important;
      font-weight: 600 !important;
  }}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  ESTADO DE SESIÓN  (modelos persistentes)
# ═══════════════════════════════════════════════════════════════════════════════
def _init_session():
    defaults = {
        "model_wat":        None,
        "scaler_wat_x":     None,
        "metrics_wat":      {},
        "model_hydrate":    None,
        "scaler_hydrate":   None,
        "metrics_hydrate":  {},
        "model_hybrid":     None,
        "scaler_hybrid_x":  None,
        "scaler_hybrid_y":  None,
        "metrics_hybrid":   {},
        "stats_model":      None,
        "logo_b64":         None,
        "models_loaded":    False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_session()


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ═══════════════════════════════════════════════════════════════════════════════
def img_to_b64(img_bytes):
    return base64.b64encode(img_bytes).decode()

def load_csv_safe(path):
    try:
        return pd.read_csv(path).dropna().drop_duplicates()
    except Exception as e:
        st.error(f"Error cargando {os.path.basename(path)}: {e}")
        return None

def save_model_artifacts(tag, model, *scalers):
    model.save(os.path.join(MODELS_DIR, f"{tag}_nn.keras"), overwrite=True)
    for i, sc in enumerate(scalers):
        if sc is not None:
            joblib.dump(sc, os.path.join(MODELS_DIR, f"{tag}_scaler_{i}.pkl"))

def load_model_artifacts(tag, n_scalers):
    mp = os.path.join(MODELS_DIR, f"{tag}_nn.keras")
    if not os.path.exists(mp):
        return None, [None]*n_scalers
    model = tf.keras.models.load_model(mp)
    scalers = []
    for i in range(n_scalers):
        sp = os.path.join(MODELS_DIR, f"{tag}_scaler_{i}.pkl")
        scalers.append(joblib.load(sp) if os.path.exists(sp) else None)
    return model, scalers

def risk_badge(status):
    if "SEGURO" in status:
        return f'<span class="badge-safe">✅ {status}</span>'
    elif "PRECAUCIÓN" in status:
        return f'<span class="badge-warn">⚠️ {status}</span>'
    else:
        return f'<span class="badge-danger">🚨 {status}</span>'

def analyse_risk(t_op, t_formation):
    margin = t_op - t_formation
    if margin < 0:
        return "CRÍTICO: Formación Activa", margin
    elif margin < 5:
        return "PRECAUCIÓN: Margen Estrecho", margin
    else:
        return "SEGURO", margin

def fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTRUCCIÓN DE MODELOS
# ═══════════════════════════════════════════════════════════════════════════════
def build_wat_model(input_shape):
    m = tf.keras.Sequential([
        tf.keras.layers.Input(shape=(input_shape,)),
        tf.keras.layers.Dense(64,  activation='swish'),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dense(128, activation='swish'),
        tf.keras.layers.Dropout(0.2),
        tf.keras.layers.Dense(32,  activation='relu'),
        tf.keras.layers.Dense(1,   activation='linear'),
    ])
    m.compile(optimizer=tf.keras.optimizers.Adam(0.0008), loss='mse', metrics=['mae'])
    return m

def build_hydrate_model():
    m = tf.keras.Sequential([
        tf.keras.layers.Dense(32, activation='softplus', input_shape=(1,)),
        tf.keras.layers.Dense(16, activation='softplus'),
        tf.keras.layers.Dense(1),
    ])
    m.compile(optimizer=tf.keras.optimizers.Adam(0.01), loss='mse')
    return m

def build_hybrid_model():
    m = tf.keras.Sequential([
        tf.keras.layers.Dense(128, activation='swish', input_shape=(2,)),
        tf.keras.layers.BatchNormalization(),
        tf.keras.layers.Dense(64,  activation='swish'),
        tf.keras.layers.Dense(32,  activation='relu'),
        tf.keras.layers.Dense(1,   activation='linear'),
    ])
    m.compile(optimizer=tf.keras.optimizers.Adam(0.0005), loss='mse')
    return m

CB_EARLY = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=30, restore_best_weights=True)


# ─── Entrenamiento WAT ───────────────────────────────────────────────────────
def train_wat(df=None, progress_cb=None):
    if df is None:
        path = os.path.join(DATA_DIR, "CSV_Parafinas.csv")
        df = load_csv_safe(path)
    if df is None:
        return False

    cols_x = ['C1-C7','C8-C15','C16-C23','C24-C30','Pressure','%Paraffin']
    col_y  = 'WAT'
    df = df[[*cols_x, col_y]].dropna()

    X, y = df[cols_x], df[col_y]
    Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.2, random_state=42)

    sc = RobustScaler()
    Xtr_s = sc.fit_transform(Xtr)
    Xte_s = sc.transform(Xte)

    model = build_wat_model(Xtr_s.shape[1])
    if progress_cb:
        progress_cb("Entrenando modelo WAT…", 0.2)
    model.fit(Xtr_s, ytr, epochs=300, batch_size=16, validation_split=0.2,
              callbacks=[CB_EARLY], verbose=0)

    yp = model.predict(Xte_s).flatten()
    r2   = r2_score(yte, yp)
    mape = mean_absolute_percentage_error(yte, yp) * 100
    rmse = np.sqrt(mean_squared_error(yte, yp))

    st.session_state.model_wat    = model
    st.session_state.scaler_wat_x = sc
    st.session_state.metrics_wat  = {"R²": r2, "MAPE (%)": mape, "RMSE (°C)": rmse}
    save_model_artifacts("wat", model, sc)
    return True


# ─── Entrenamiento Hidratos puros ────────────────────────────────────────────
def train_hydrate(df=None, progress_cb=None):
    if df is None:
        path = os.path.join(DATA_DIR, "Base_de_Datos_Doctorado_1_co.csv")
        df = load_csv_safe(path)
    if df is None:
        return False

    df.columns = df.columns.str.strip()
    T_col = [c for c in df.columns if 'T' in c.upper()][0]
    P_col = [c for c in df.columns if 'P' in c.upper()][0]

    X = df[[T_col]].values
    y = np.log(df[P_col].values)

    Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.2, random_state=42)
    sc = StandardScaler()
    Xtr_s = sc.fit_transform(Xtr)
    Xte_s = sc.transform(Xte)

    model = build_hydrate_model()
    if progress_cb:
        progress_cb("Entrenando red de hidratos…", 0.4)
    model.fit(Xtr_s, ytr, epochs=300, batch_size=8, validation_split=0.1, verbose=0)

    yp_log = model.predict(Xte_s).flatten()
    yp = np.exp(yp_log)
    yt = np.exp(yte)
    r2   = r2_score(yt, yp)
    mape = mean_absolute_percentage_error(yt, yp) * 100

    st.session_state.model_hydrate   = model
    st.session_state.scaler_hydrate  = sc
    st.session_state.metrics_hydrate = {"R²": r2, "MAPE (%)": mape}
    save_model_artifacts("hydrate", model, sc)
    return True


# ─── Entrenamiento Híbrido (CH4 + Metanol) ───────────────────────────────────
def train_hybrid(df=None, progress_cb=None):
    if df is None:
        path = os.path.join(DATA_DIR, "CH4_Metanol__P_T_XCH4O.csv")
        df = load_csv_safe(path)
    if df is None:
        return False

    df.columns = df.columns.str.strip()
    df['log_P'] = np.log(df['P(kPa)'])

    X_in  = df[['xCH4O', 'log_P']].values
    y_out = df['T(K)'].values.reshape(-1, 1)

    Xtr, Xte, ytr, yte = train_test_split(X_in, y_out, test_size=0.2, random_state=42)
    sc_x = RobustScaler(); sc_y = RobustScaler()
    Xtr_s = sc_x.fit_transform(Xtr)
    Xte_s = sc_x.transform(Xte)
    ytr_s = sc_y.fit_transform(ytr)

    # Stats model
    stats_df = df.copy()
    stats_model = smf.ols('Q("T(K)") ~ xCH4O + log_P', data=stats_df).fit()

    model = build_hybrid_model()
    if progress_cb:
        progress_cb("Entrenando modelo híbrido…", 0.6)
    model.fit(Xtr_s, ytr_s, epochs=600, batch_size=8, validation_split=0.2,
              callbacks=[CB_EARLY], verbose=0)

    yp_s = model.predict(Xte_s)
    yp   = sc_y.inverse_transform(yp_s).flatten()
    yt   = yte.flatten()
    r2   = r2_score(yt, yp)
    mape = mean_absolute_percentage_error(yt, yp) * 100

    st.session_state.model_hybrid    = model
    st.session_state.scaler_hybrid_x = sc_x
    st.session_state.scaler_hybrid_y = sc_y
    st.session_state.stats_model     = stats_model
    st.session_state.metrics_hybrid  = {"R²": r2, "MAPE (%)": mape,
                                         "R² Estadístico": stats_model.rsquared_adj}
    save_model_artifacts("hybrid", model, sc_x, sc_y)
    return True


# ─── Carga de modelos guardados ──────────────────────────────────────────────
def try_load_saved_models():
    if st.session_state.models_loaded:
        return
    m, sc = load_model_artifacts("wat", 1)
    if m:
        st.session_state.model_wat    = m
        st.session_state.scaler_wat_x = sc[0]

    m, sc = load_model_artifacts("hydrate", 1)
    if m:
        st.session_state.model_hydrate  = m
        st.session_state.scaler_hydrate = sc[0]

    m, sc = load_model_artifacts("hybrid", 2)
    if m:
        st.session_state.model_hybrid    = m
        st.session_state.scaler_hybrid_x = sc[0]
        st.session_state.scaler_hybrid_y = sc[1]

    st.session_state.models_loaded = True

try_load_saved_models()


# ═══════════════════════════════════════════════════════════════════════════════
#  GRÁFICAS
# ═══════════════════════════════════════════════════════════════════════════════
def plot_hydrate_envelope(T_op_K, P_op_kPa, inhibitor_conc=0.0, has_nacl=False,
                          xNaCl=0.0, logo_b64=None):
    model  = st.session_state.model_hydrate
    sc     = st.session_state.scaler_hydrate
    hybrid = st.session_state.model_hybrid
    sc_hx  = st.session_state.scaler_hybrid_x
    sc_hy  = st.session_state.scaler_hybrid_y

    fig, ax = plt.subplots(figsize=(9, 6))
    fig.patch.set_facecolor('#fafafa')
    ax.set_facecolor('#fafafa')

    T_range = np.linspace(265, 310, 200).reshape(-1, 1)

    # Curva base (CH4 puro)
    if model and sc:
        T_s = sc.transform(T_range)
        P_base = np.exp(model.predict(T_s).flatten())
        ax.plot(T_range.flatten(), P_base/1000, color=IMP_GREEN, lw=2.5,
                label='CH₄ puro (IA)', zorder=4)
        ax.fill_between(T_range.flatten(), P_base/1000,
                         color=IMP_DANGER, alpha=0.12, zorder=2)

    # Curva con inhibidor
    if hybrid and sc_hx and sc_hy and inhibitor_conc > 0:
        log_Ps = np.log(np.clip(P_base, 1e-6, None))
        X_hyb  = np.column_stack([np.full(200, inhibitor_conc), log_Ps])
        X_s    = sc_hx.transform(X_hyb)
        T_inh  = sc_hy.inverse_transform(hybrid.predict(X_s)).flatten()
        ax.plot(T_inh, P_base/1000, '--', color=IMP_GOLD, lw=2.5,
                label=f'Con metanol ({inhibitor_conc:.2f} fm)', zorder=4)

    # Punto de operación
    ax.scatter([T_op_K], [P_op_kPa/1000], color='royalblue', s=180, zorder=10,
               edgecolors='white', linewidths=2, label=f'Condición de operación')

    # Anotación
    ax.annotate(f'T={T_op_K:.1f} K\nP={P_op_kPa:.0f} kPa',
                xy=(T_op_K, P_op_kPa/1000),
                xytext=(T_op_K+1.5, P_op_kPa/1000*1.2),
                fontsize=8.5, fontfamily='monospace',
                arrowprops=dict(arrowstyle='->', color='royalblue', lw=1.5),
                color='royalblue', fontweight='600')

    ax.set_xlabel("Temperatura (K)", fontsize=11, fontfamily='sans-serif')
    ax.set_ylabel("Presión (MPa)", fontsize=11, fontfamily='sans-serif')
    ax.set_title("Envolvente de Fase — Hidratos de Gas", fontsize=13, fontweight='bold',
                 color=IMP_GREEN, pad=14)
    ax.legend(fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.3, linestyle='--')

    # Zonas
    ax.axvspan(ax.get_xlim()[0] if ax.get_xlim()[0] > 265 else 265,
               min(T_op_K, T_range.max()), alpha=0.06, color=IMP_DANGER)

    fig.tight_layout()
    return fig


def plot_wat_envelope(T_op_C, P_op_bar, c1c7, c8c15, c16c23, c24c30, pct_par):
    model = st.session_state.model_wat
    sc    = st.session_state.scaler_wat_x

    fig, ax = plt.subplots(figsize=(9, 6))
    fig.patch.set_facecolor('#fafafa')
    ax.set_facecolor('#fafafa')

    if model and sc:
        pressures = np.linspace(2000, 15000, 100)
        wats = []
        for p in pressures:
            X_in = np.array([[c1c7, c8c15, c16c23, c24c30, p, pct_par]])
            X_s  = sc.transform(X_in)
            wats.append(model.predict(X_s, verbose=0).flatten()[0])
        wats = np.array(wats)

        ax.plot(pressures/100, wats, color=IMP_GREEN, lw=2.5, label='Curva WAT (IA)', zorder=4)
        ax.fill_between(pressures/100, wats,
                        np.full_like(wats, wats.max()+5),
                        color=IMP_DANGER, alpha=0.12, label='Zona de Riesgo', zorder=2)
        ax.fill_between(pressures/100, wats,
                        np.full_like(wats, wats.min()-5),
                        color=IMP_LIGHT, alpha=0.7, zorder=2)

    ax.scatter([P_op_bar], [T_op_C], color='royalblue', s=180, zorder=10,
               edgecolors='white', linewidths=2, label='Condición de operación')

    ax.set_xlabel("Presión (bar)", fontsize=11)
    ax.set_ylabel("WAT (°C)", fontsize=11)
    ax.set_title("Temperatura de Aparición de Cera (WAT)", fontsize=13,
                 fontweight='bold', color=IMP_GREEN, pad=14)
    ax.legend(fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.3, linestyle='--')
    fig.tight_layout()
    return fig


def plot_inhibitor_sensitivity(P_kPa):
    model  = st.session_state.model_hybrid
    sc_hx  = st.session_state.scaler_hybrid_x
    sc_hy  = st.session_state.scaler_hybrid_y
    if not (model and sc_hx and sc_hy):
        return None

    concs   = np.linspace(0, 0.5, 60)
    log_P   = np.log(P_kPa)
    X_in    = np.column_stack([concs, np.full(60, log_P)])
    X_s     = sc_hx.transform(X_in)
    T_pred  = sc_hy.inverse_transform(model.predict(X_s)).flatten()

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(concs*100, T_pred - 273.15, color=IMP_GOLD, lw=2.5)
    ax.fill_between(concs*100, T_pred-273.15, T_pred.min()-273.15-2,
                    color=IMP_GOLD, alpha=0.15)
    ax.set_xlabel("Concentración de Metanol (%fm × 100)", fontsize=10)
    ax.set_ylabel("T de Formación de Hidratos (°C)", fontsize=10)
    ax.set_title("Análisis de Sensibilidad — Inhibidor Termodinámico",
                 fontsize=12, fontweight='bold', color=IMP_GREEN)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.annotate("Mayor inhibición →", xy=(40, T_pred[-1]-273.15),
                fontsize=9, color=IMP_GREEN, fontweight='600')
    fig.tight_layout()
    return fig


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORTACIÓN EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def export_excel(df_in, df_out, module_name):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Portada
        pd.DataFrame({
            "Software": ["Data & Flow Predictor"],
            "Instituto": ["Instituto Mexicano del Petróleo"],
            "Módulo": [module_name],
            "Fecha": [pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")],
        }).to_excel(writer, sheet_name="Portada", index=False)

        df_in.to_excel(writer,  sheet_name="Datos de Entrada", index=False)
        df_out.to_excel(writer, sheet_name="Predicciones",     index=False)

        # Formato
        wb = writer.book
        from openpyxl.styles import PatternFill, Font, Alignment
        green_fill = PatternFill("solid", fgColor="006847")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill   = green_fill
                cell.font   = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18

    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⛽ Data & Flow")
    st.markdown("**Predictor v2.0**")
    st.divider()

    logo_file = st.file_uploader("📎 Logo del Posgrado (PNG/JPG)", type=["png","jpg","jpeg"])
    if logo_file:
        b64 = img_to_b64(logo_file.read())
        st.session_state.logo_b64 = b64
        st.image(f"data:image/png;base64,{b64}", width=160)
    elif st.session_state.logo_b64:
        st.image(f"data:image/png;base64,{st.session_state.logo_b64}", width=160)

    st.divider()
    page = st.radio("**Navegación**", [
        "🏠 Dashboard",
        "💧 Simulador Hidratos",
        "🕯️ Simulador Parafinas",
        "🔁 Re-Entrenamiento",
        "📄 Reportes",
    ])
    st.divider()

    # Estado de modelos
    st.markdown("**Estado de Modelos**")
    def dot(ok): return "🟢" if ok else "🔴"
    st.markdown(f"{dot(st.session_state.model_wat)}  WAT (Parafinas)")
    st.markdown(f"{dot(st.session_state.model_hydrate)}  Hidratos CH₄ Puro")
    st.markdown(f"{dot(st.session_state.model_hybrid)}  Híbrido CH₄+MeOH")

    st.divider()
    st.markdown('<p style="font-size:0.7rem;opacity:0.7">© IMP · Posgrado en Ingeniería de Hidrocarburos</p>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  HEADER GLOBAL
# ═══════════════════════════════════════════════════════════════════════════════
logo_html = ""
if st.session_state.logo_b64:
    logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" height="54" style="border-radius:6px;"/>'

st.markdown(f"""
<div class="imp-header">
  {logo_html}
  <div>
    <h1>Data &amp; Flow Predictor</h1>
    <span>Instituto Mexicano del Petróleo · Aseguramiento de Flujo · Hidratos &amp; Parafinas</span>
  </div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PÁGINA 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if page == "🏠 Dashboard":
    st.markdown('<div class="sec-title">📊 Panel de Control del Sistema</div>', unsafe_allow_html=True)

    k1, k2, k3, k4 = st.columns(4)
    models_ok = sum([
        st.session_state.model_wat is not None,
        st.session_state.model_hydrate is not None,
        st.session_state.model_hybrid is not None,
    ])

    with k1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-val">{models_ok}/3</div><div class="kpi-lbl">Modelos Activos</div></div>', unsafe_allow_html=True)
    with k2:
        r2_w = st.session_state.metrics_wat.get("R²", "—")
        val  = f"{r2_w:.4f}" if isinstance(r2_w, float) else r2_w
        st.markdown(f'<div class="kpi-card gold"><div class="kpi-val">{val}</div><div class="kpi-lbl">R² Modelo WAT</div></div>', unsafe_allow_html=True)
    with k3:
        r2_h = st.session_state.metrics_hybrid.get("R²", "—")
        val  = f"{r2_h:.4f}" if isinstance(r2_h, float) else r2_h
        st.markdown(f'<div class="kpi-card"><div class="kpi-val">{val}</div><div class="kpi-lbl">R² Mod. Hidratos</div></div>', unsafe_allow_html=True)
    with k4:
        mape = st.session_state.metrics_hydrate.get("MAPE (%)", "—")
        val  = f"{mape:.2f}%" if isinstance(mape, float) else mape
        st.markdown(f'<div class="kpi-card warn"><div class="kpi-val">{val}</div><div class="kpi-lbl">MAPE CH₄ Puro</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-title">🗄️ Datos de Entrenamiento Disponibles</div>', unsafe_allow_html=True)

    csv_files = {
        "CSV_Parafinas.csv":                    ("WAT / Parafinas",      ["C1-C7","C8-C15","C16-C23","C24-C30","Pressure","%Paraffin","WAT"]),
        "Base_de_Datos_Doctorado_1_co.csv":     ("Hidratos CH₄ puro",    ["T(K)","P(kPa)"]),
        "CH4_Metanol__P_T_XCH4O.csv":           ("CH₄ + Metanol",        ["P(kPa)","T(K)","xCH4O"]),
        "Metanol_y_NaCl_P_t__xch4_y_xNaCl.csv":("CH₄ + MeOH + NaCl",   ["P(kPa)","T(K)","xClNa","xCH4O"]),
        "CH4_con_xCH4_.csv":                    ("CH₄ fracción molar",   ["P(kPa)","T(K)","xCH4"]),
        "Base_de_Datos_Doctorado_1_componente.csv": ("1 Componente IMP", ["T (K)","P (Mpa)"]),
    }

    rows = []
    for fn, (desc, cols) in csv_files.items():
        fpath = os.path.join(DATA_DIR, fn)
        if os.path.exists(fpath):
            df_tmp = pd.read_csv(fpath)
            rows.append({"Archivo": fn, "Descripción": desc, "Muestras": len(df_tmp), "Columnas": ", ".join(df_tmp.columns.tolist())})
        else:
            rows.append({"Archivo": fn, "Descripción": desc, "Muestras": "—", "Columnas": "No encontrado"})

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown('<div class="sec-title">🚀 Entrenamiento Inicial (Todos los Modelos)</div>', unsafe_allow_html=True)
    st.info("Si es la primera vez o deseas re-entrenar todos los modelos con los datos originales, haz clic abajo.")

    if st.button("⚡ Entrenar / Recargar Todos los Modelos"):
        prog = st.progress(0)
        msg  = st.empty()

        def cb(text, val):
            msg.markdown(f"**{text}**")
            prog.progress(val)

        with st.spinner("Entrenando…"):
            cb("Entrenando modelo WAT (Parafinas)…", 0.1)
            ok1 = train_wat(progress_cb=cb)
            cb("Entrenando red de Hidratos CH₄ puro…", 0.45)
            ok2 = train_hydrate(progress_cb=cb)
            cb("Entrenando modelo híbrido CH₄ + Metanol…", 0.7)
            ok3 = train_hybrid(progress_cb=cb)
            prog.progress(1.0)
            msg.empty()

        if ok1 and ok2 and ok3:
            st.success("✅ Todos los modelos entrenados y guardados exitosamente.")
        else:
            st.warning("⚠️ Algún archivo CSV no fue localizado. Verifica la carpeta `data/`.")
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  PÁGINA 2: SIMULADOR HIDRATOS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "💧 Simulador Hidratos":
    st.markdown('<div class="sec-title">💧 Simulador de Hidratos de Gas</div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["🔢 Predicción Individual", "📂 Carga Batch", "📈 Análisis de Inhibidor"])

    # ── Tab 1: Individual ──────────────────────────────────────────────────
    with tab1:
        col_in, col_out = st.columns([1, 1.8])

        with col_in:
            st.markdown("**Condiciones de Operación**")
            T_op = st.number_input("Temperatura de Operación (K)", 265.0, 320.0, 285.0, 0.5)
            P_op = st.number_input("Presión de Operación (kPa)",   500.0, 50000.0, 8000.0, 100.0)

            st.markdown("**Inhibidor Termodinámico**")
            use_inhibitor = st.checkbox("Incluir Metanol")
            x_meoh = 0.0
            if use_inhibitor:
                x_meoh = st.slider("Fracción Molar MeOH (xCH₄O)", 0.0, 0.5, 0.1, 0.01)

            st.markdown("**Electrolito (NaCl)**")
            use_nacl = st.checkbox("Incluir NaCl")
            x_nacl = 0.0
            if use_nacl:
                x_nacl = st.slider("Fracción Molar NaCl", 0.0, 0.10, 0.02, 0.005)

            predict_btn = st.button("🔬 Predecir Temperatura de Formación", use_container_width=True)

        with col_out:
            if predict_btn:
                model_h  = st.session_state.model_hydrate
                sc_h     = st.session_state.scaler_hydrate
                model_hy = st.session_state.model_hybrid
                sc_hx    = st.session_state.scaler_hybrid_x
                sc_hy    = st.session_state.scaler_hybrid_y

                if not (model_h and sc_h):
                    st.error("❌ Modelo de hidratos no entrenado. Ve al Dashboard y entrena primero.")
                else:
                    # Predicción base (CH4 puro)
                    T_s  = sc_h.transform([[T_op]])
                    P_pred_base = np.exp(model_h.predict(T_s, verbose=0).flatten()[0])

                    # Predicción con inhibidor
                    T_form_K = T_op  # Por defecto
                    if use_inhibitor and model_hy and sc_hx and sc_hy:
                        log_P = np.log(P_op)
                        X_in  = np.array([[x_meoh, log_P]])
                        X_s   = sc_hx.transform(X_in)
                        T_form_K = sc_hy.inverse_transform(model_hy.predict(X_s, verbose=0)).flatten()[0]
                    else:
                        # Inferir T de formación de la curva base dado P_op
                        T_range = np.linspace(265, 315, 500).reshape(-1,1)
                        T_s2 = sc_h.transform(T_range)
                        P_curve = np.exp(model_h.predict(T_s2, verbose=0).flatten())
                        idx = np.argmin(np.abs(P_curve - P_op))
                        T_form_K = T_range.flatten()[idx]

                    status, margin = analyse_risk(T_op, T_form_K)

                    st.markdown("### Resultados")
                    m1, m2, m3 = st.columns(3)
                    m1.metric("T Formación (K)", f"{T_form_K:.2f}")
                    m2.metric("T Operación (K)", f"{T_op:.2f}")
                    m3.metric("Margen ΔT (K)", f"{margin:.2f}", delta=f"{margin:.2f} K")

                    st.markdown(f"**Estado del Sistema:** {risk_badge(status)}", unsafe_allow_html=True)

                    mape = st.session_state.metrics_hydrate.get("MAPE (%)", None)
                    diag_text = f"""
**Diagnóstico Automático:**
La red neuronal predice una temperatura de formación de hidratos de **{T_form_K:.2f} K** ({T_form_K-273.15:.2f} °C)
a las condiciones de operación P = {P_op:.0f} kPa.
El margen de seguridad es de **ΔT = {margin:.2f} K**.
{'El metanol desplaza la curva de formación en aprox. ' + str(abs(margin))[:4] + ' K.' if use_inhibitor else ''}
{"MAPE del modelo: " + f"{mape:.2f}%" if mape else ""}
"""
                    with st.expander("📋 Ver Diagnóstico Completo", expanded=True):
                        st.markdown(diag_text)

                    # Guardar para reporte
                    st.session_state["last_hyd_result"] = {
                        "T_op": T_op, "P_op": P_op, "T_form": T_form_K,
                        "margin": margin, "status": status,
                        "x_meoh": x_meoh, "x_nacl": x_nacl,
                    }

            # Gráfica siempre visible si hay modelo
            if st.session_state.model_hydrate:
                T_plot = st.session_state.get("last_hyd_result", {}).get("T_op", 285.0)
                P_plot = st.session_state.get("last_hyd_result", {}).get("P_op", 8000.0)
                xm     = st.session_state.get("last_hyd_result", {}).get("x_meoh", 0.0)
                fig = plot_hydrate_envelope(T_plot, P_plot, inhibitor_conc=xm)
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)
            else:
                st.info("Entrena los modelos desde el Dashboard para ver la envolvente.")

    # ── Tab 2: Batch ───────────────────────────────────────────────────────
    with tab2:
        st.markdown("**Sube un CSV con columnas:** `T(K)`, `P(kPa)`, `xCH4O` (opcional)")
        batch_file = st.file_uploader("Archivo CSV Batch", type="csv", key="hyd_batch")

        if batch_file:
            df_b = pd.read_csv(batch_file).dropna()
            st.dataframe(df_b.head(), use_container_width=True)

            if st.button("🔄 Procesar Batch"):
                model_h  = st.session_state.model_hydrate
                sc_h     = st.session_state.scaler_hydrate
                model_hy = st.session_state.model_hybrid
                sc_hx    = st.session_state.scaler_hybrid_x
                sc_hy    = st.session_state.scaler_hybrid_y

                if not (model_h and sc_h):
                    st.error("Modelos no entrenados.")
                else:
                    results = []
                    for _, row in df_b.iterrows():
                        T_r = row.get('T(K)', 285.0)
                        P_r = row.get('P(kPa)', 8000.0)
                        xm  = row.get('xCH4O', 0.0)

                        # T formación
                        T_range = np.linspace(265, 315, 500).reshape(-1,1)
                        T_s2 = sc_h.transform(T_range)
                        P_cur = np.exp(model_h.predict(T_s2, verbose=0).flatten())
                        idx   = np.argmin(np.abs(P_cur - P_r))
                        T_f   = T_range.flatten()[idx]

                        if xm > 0 and model_hy and sc_hx and sc_hy:
                            log_P  = np.log(max(P_r, 1e-6))
                            X_in   = np.array([[xm, log_P]])
                            X_s    = sc_hx.transform(X_in)
                            T_f    = sc_hy.inverse_transform(model_hy.predict(X_s, verbose=0)).flatten()[0]

                        status, margin = analyse_risk(T_r, T_f)
                        results.append({"T_Op (K)": T_r, "P_Op (kPa)": P_r, "xCH4O": xm,
                                         "T_Form (K)": round(T_f,2), "ΔT (K)": round(margin,2), "Estado": status})

                    df_res = pd.DataFrame(results)
                    st.dataframe(df_res, use_container_width=True)

                    xlsx = export_excel(df_b, df_res, "Hidratos Batch")
                    st.download_button("📥 Descargar Excel", xlsx,
                                       file_name="Hidratos_Batch.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Tab 3: Análisis de Inhibidor ───────────────────────────────────────
    with tab3:
        st.markdown("**Sensibilidad de la temperatura de formación ante variaciones en la concentración de metanol.**")
        P_sens = st.number_input("Presión fija (kPa)", 500.0, 50000.0, 10000.0, 500.0)
        if st.button("📊 Generar Análisis de Sensibilidad"):
            if not st.session_state.model_hybrid:
                st.error("Modelo híbrido no entrenado.")
            else:
                fig = plot_inhibitor_sensitivity(P_sens)
                if fig:
                    st.pyplot(fig, use_container_width=True)
                    plt.close(fig)


# ═══════════════════════════════════════════════════════════════════════════════
#  PÁGINA 3: SIMULADOR PARAFINAS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🕯️ Simulador Parafinas":
    st.markdown('<div class="sec-title">🕯️ Simulador de Temperatura de Aparición de Cera (WAT)</div>', unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["🔢 Predicción Individual", "📂 Carga Batch"])

    with tab1:
        col_in, col_out = st.columns([1, 1.8])

        with col_in:
            st.markdown("**Composición de la Muestra (%)**")
            c1c7  = st.number_input("C₁–C₇ (%)",   0.0, 100.0, 9.79, 0.1)
            c8c15 = st.number_input("C₈–C₁₅ (%)",  0.0, 100.0, 62.89, 0.1)
            c16c23= st.number_input("C₁₆–C₂₃ (%)", 0.0, 100.0, 22.36, 0.1)
            c24c30= st.number_input("C₂₄–C₃₀ (%)", 0.0, 100.0, 4.96, 0.1)
            total = c1c7+c8c15+c16c23+c24c30
            if abs(total-100) > 5:
                st.warning(f"Composición suma {total:.1f}% (idealmente 100%)")

            st.markdown("**Condiciones del Sistema**")
            pres_par = st.number_input("Presión (kPa)", 500.0, 20000.0, 8000.0, 100.0)
            pct_par  = st.number_input("% Parafina", 0.0, 100.0, 26.6, 0.1)
            T_op_par = st.number_input("Temperatura de Operación (°C)", 0.0, 120.0, 50.0, 0.5)

            btn_wat = st.button("🔬 Predecir WAT", use_container_width=True)

        with col_out:
            if btn_wat:
                model_w = st.session_state.model_wat
                sc_w    = st.session_state.scaler_wat_x

                if not (model_w and sc_w):
                    st.error("❌ Modelo WAT no entrenado. Ve al Dashboard.")
                else:
                    X_in = np.array([[c1c7, c8c15, c16c23, c24c30, pres_par, pct_par]])
                    X_s  = sc_w.transform(X_in)
                    wat_pred = model_w.predict(X_s, verbose=0).flatten()[0]
                    margin_wat = T_op_par - wat_pred
                    status_wat, _ = analyse_risk(T_op_par + 273.15, wat_pred + 273.15)

                    w1, w2, w3 = st.columns(3)
                    w1.metric("WAT Predicha (°C)", f"{wat_pred:.2f}")
                    w2.metric("T Operación (°C)", f"{T_op_par:.2f}")
                    w3.metric("Margen ΔT (°C)", f"{margin_wat:.2f}", delta=f"{margin_wat:.2f}°C")

                    st.markdown(f"**Estado:** {risk_badge(status_wat)}", unsafe_allow_html=True)

                    mape_w = st.session_state.metrics_wat.get("MAPE (%)", None)
                    rmse_w = st.session_state.metrics_wat.get("RMSE (°C)", None)
                    with st.expander("📋 Diagnóstico Técnico", expanded=True):
                        st.markdown(f"""
**WAT Predicha:** {wat_pred:.2f} °C  
**Margen de Seguridad:** ΔT = {margin_wat:.2f} °C  
**Composición:** C₁–C₇={c1c7}%, C₈–C₁₅={c8c15}%, C₁₆–C₂₃={c16c23}%, C₂₄–C₃₀={c24c30}%  
**% Parafina:** {pct_par}%  
{"**MAPE del modelo:** " + f"{mape_w:.2f}%" if mape_w else ""}  
{"**RMSE:** " + f"{rmse_w:.2f} °C" if rmse_w else ""}  

{"⚠️ **La temperatura de operación es INFERIOR a la WAT**. Riesgo de deposición de parafinas." if margin_wat < 0 else "✅ Operando por encima de WAT. Sin riesgo de deposición."}
""")
                    st.session_state["last_wat_result"] = {
                        "wat": wat_pred, "T_op": T_op_par, "pres": pres_par,
                        "c1c7": c1c7, "c8c15": c8c15, "c16c23": c16c23,
                        "c24c30": c24c30, "pct_par": pct_par,
                    }

            if st.session_state.model_wat:
                r = st.session_state.get("last_wat_result", {})
                fig_w = plot_wat_envelope(
                    r.get("T_op",50), r.get("pres",8000)/100,
                    r.get("c1c7",9.79), r.get("c8c15",62.89),
                    r.get("c16c23",22.36), r.get("c24c30",4.96),
                    r.get("pct_par",26.6)
                )
                st.pyplot(fig_w, use_container_width=True)
                plt.close(fig_w)
            else:
                st.info("Entrena los modelos desde el Dashboard.")

    with tab2:
        st.markdown("**Columnas requeridas:** `C1-C7`, `C8-C15`, `C16-C23`, `C24-C30`, `Pressure`, `%Paraffin`")
        batch_par = st.file_uploader("CSV de Composiciones", type="csv", key="par_batch")

        if batch_par:
            df_pb = pd.read_csv(batch_par).dropna()
            st.dataframe(df_pb.head(), use_container_width=True)

            if st.button("🔄 Predecir Batch WAT"):
                model_w = st.session_state.model_wat
                sc_w    = st.session_state.scaler_wat_x
                if not (model_w and sc_w):
                    st.error("Modelo WAT no entrenado.")
                else:
                    cols = ['C1-C7','C8-C15','C16-C23','C24-C30','Pressure','%Paraffin']
                    missing = [c for c in cols if c not in df_pb.columns]
                    if missing:
                        st.error(f"Columnas faltantes: {missing}")
                    else:
                        X_b = df_pb[cols].values
                        X_bs = sc_w.transform(X_b)
                        wats = model_w.predict(X_bs, verbose=0).flatten()
                        df_res_p = df_pb.copy()
                        df_res_p['WAT_Predicha (°C)'] = np.round(wats, 2)
                        st.dataframe(df_res_p, use_container_width=True)
                        xlsx = export_excel(df_pb, df_res_p, "WAT Parafinas Batch")
                        st.download_button("📥 Descargar Excel", xlsx,
                                           file_name="WAT_Batch.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
#  PÁGINA 4: RE-ENTRENAMIENTO
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔁 Re-Entrenamiento":
    st.markdown('<div class="sec-title">🔁 Módulo Evolutivo de Re-Entrenamiento</div>', unsafe_allow_html=True)
    st.info("Sube nuevos datos experimentales para actualizar las redes neuronales. Los modelos se guardan localmente y persisten entre sesiones.")

    module = st.selectbox("Módulo a re-entrenar", [
        "WAT (Parafinas)",
        "Hidratos CH₄ Puro",
        "Híbrido CH₄ + Metanol",
    ])

    col_fmt = {
        "WAT (Parafinas)":          "Sample, Pressure, %Paraffin, C1-C7, C8-C15, C16-C23, C24-C30, **WAT**",
        "Hidratos CH₄ Puro":        "**T(K)**, **P(kPa)**",
        "Híbrido CH₄ + Metanol":    "**P(kPa)**, **T(K)**, **xCH4O**",
    }
    st.markdown(f"**Formato requerido:** `{col_fmt[module]}`")

    new_csv = st.file_uploader("Subir CSV con Nuevos Datos", type="csv", key="retrain_csv")

    if new_csv:
        df_new = pd.read_csv(new_csv).dropna()
        st.markdown(f"**{len(df_new)} muestras cargadas.** Vista previa:")
        st.dataframe(df_new.head(8), use_container_width=True)

        if st.button("⚡ Re-Entrenar Modelo", use_container_width=True):
            prog2 = st.progress(0)
            msg2  = st.empty()
            def cb2(t, v): msg2.markdown(f"**{t}**"); prog2.progress(v)

            with st.spinner("Entrenando…"):
                if module == "WAT (Parafinas)":
                    ok = train_wat(df=df_new, progress_cb=cb2)
                    metrics = st.session_state.metrics_wat
                elif module == "Hidratos CH₄ Puro":
                    ok = train_hydrate(df=df_new, progress_cb=cb2)
                    metrics = st.session_state.metrics_hydrate
                else:
                    ok = train_hybrid(df=df_new, progress_cb=cb2)
                    metrics = st.session_state.metrics_hybrid

            prog2.progress(1.0); msg2.empty()

            if ok:
                st.success("✅ Modelo actualizado y guardado.")
                st.markdown("**Nuevas Métricas:**")
                cols_m = st.columns(len(metrics))
                for (k, v), col in zip(metrics.items(), cols_m):
                    col.metric(k, f"{v:.4f}" if isinstance(v, float) else v)
            else:
                st.error("Error durante el entrenamiento. Revisa el formato del CSV.")

    st.divider()
    st.markdown("**📁 Ruta de Almacenamiento de Modelos:**")
    st.code(MODELS_DIR)
    saved = os.listdir(MODELS_DIR) if os.path.exists(MODELS_DIR) else []
    if saved:
        st.markdown(f"Archivos guardados: `{'`, `'.join(saved)}`")
    else:
        st.markdown("_No hay modelos guardados aún._")


# ═══════════════════════════════════════════════════════════════════════════════
#  PÁGINA 5: REPORTES
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📄 Reportes":
    st.markdown('<div class="sec-title">📄 Generación de Reportes de Ingeniería</div>', unsafe_allow_html=True)

    rep_type = st.radio("Tipo de Reporte", ["Hidratos", "Parafinas (WAT)"], horizontal=True)

    if rep_type == "Hidratos":
        r = st.session_state.get("last_hyd_result")
        if not r:
            st.warning("Primero realiza una predicción en el Simulador de Hidratos.")
        else:
            st.markdown("### Vista Previa del Reporte")
            c1, c2 = st.columns(2)
            c1.metric("T Operación (K)",  f"{r['T_op']:.2f}")
            c1.metric("P Operación (kPa)",f"{r['P_op']:.0f}")
            c2.metric("T Formación (K)",  f"{r['T_form']:.2f}")
            c2.metric("ΔT Margen (K)",    f"{r['margin']:.2f}")
            st.markdown(f"**Estado:** {risk_badge(r['status'])}", unsafe_allow_html=True)

            fig_rep = plot_hydrate_envelope(r['T_op'], r['P_op'], r.get('x_meoh',0))
            st.pyplot(fig_rep, use_container_width=True)

            # Excel
            df_entry = pd.DataFrame([{
                "T_Op (K)": r['T_op'], "P_Op (kPa)": r['P_op'],
                "xCH4O": r.get('x_meoh',0), "xNaCl": r.get('x_nacl',0)
            }])
            df_pred = pd.DataFrame([{
                "T_Formación (K)": r['T_form'],
                "ΔT (K)": r['margin'],
                "Estado": r['status'],
                "MAPE (%)": st.session_state.metrics_hydrate.get('MAPE (%)', 'N/A'),
            }])
            xlsx_rep = export_excel(df_entry, df_pred, "Hidratos Individual")
            st.download_button(
                "📥 Descargar Reporte Excel",
                xlsx_rep,
                file_name="Reporte_Hidratos_IMP.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # PNG de la gráfica
            png_bytes = fig_to_bytes(fig_rep)
            st.download_button(
                "🖼️ Descargar Gráfica (PNG)",
                png_bytes,
                file_name="Envolvente_Hidratos.png",
                mime="image/png",
                use_container_width=True,
            )
            plt.close(fig_rep)

    else:  # WAT
        r = st.session_state.get("last_wat_result")
        if not r:
            st.warning("Primero realiza una predicción en el Simulador de Parafinas.")
        else:
            st.markdown("### Vista Previa del Reporte")
            c1, c2 = st.columns(2)
            c1.metric("WAT Predicha (°C)",   f"{r['wat']:.2f}")
            c1.metric("T Operación (°C)",     f"{r['T_op']:.2f}")
            c2.metric("Presión (kPa)",         f"{r['pres']:.0f}")
            c2.metric("% Parafina",            f"{r['pct_par']:.1f}")

            fig_wrep = plot_wat_envelope(
                r['T_op'], r['pres']/100,
                r['c1c7'], r['c8c15'], r['c16c23'], r['c24c30'], r['pct_par']
            )
            st.pyplot(fig_wrep, use_container_width=True)

            df_we = pd.DataFrame([{
                "C1-C7 (%)": r['c1c7'], "C8-C15 (%)": r['c8c15'],
                "C16-C23 (%)": r['c16c23'], "C24-C30 (%)": r['c24c30'],
                "Presión (kPa)": r['pres'], "% Parafina": r['pct_par'],
                "T_Op (°C)": r['T_op'],
            }])
            df_wp = pd.DataFrame([{
                "WAT Predicha (°C)": r['wat'],
                "ΔT (°C)": r['T_op'] - r['wat'],
                "MAPE (%)": st.session_state.metrics_wat.get('MAPE (%)', 'N/A'),
                "RMSE (°C)": st.session_state.metrics_wat.get('RMSE (°C)', 'N/A'),
            }])
            xlsx_wrep = export_excel(df_we, df_wp, "WAT Parafinas")
            st.download_button(
                "📥 Descargar Reporte Excel",
                xlsx_wrep,
                file_name="Reporte_WAT_IMP.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            png_w = fig_to_bytes(fig_wrep)
            st.download_button(
                "🖼️ Descargar Gráfica (PNG)",
                png_w,
                file_name="Curva_WAT.png",
                mime="image/png",
                use_container_width=True,
            )
            plt.close(fig_wrep)
